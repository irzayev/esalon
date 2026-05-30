"""Export tabular reports to Excel (.xlsx)."""
from __future__ import annotations

from io import BytesIO
from typing import Any, Sequence

from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models.settings import Settings


def format_money(value: Any) -> str:
    try:
        v = float(value or 0)
    except (TypeError, ValueError):
        v = 0.0
    cur = Settings.get().default_currency or "AZN"
    sym = "₼" if cur.upper() in ("AZN", "₼") else cur
    return f"{v:,.2f} {sym}".replace(",", " ")


def _auto_column_width(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1] or "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 48)


def build_excel_workbook(
    sheets: Sequence[dict[str, Any]],
) -> bytes:
    """Build xlsx bytes. Each sheet: name, headers, rows, optional summary_rows."""
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E293B")

    for spec in sheets:
        ws = wb.create_sheet(title=str(spec.get("name", "Sheet"))[:31])
        headers = list(spec.get("headers") or [])
        rows = list(spec.get("rows") or [])
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in rows:
            ws.append([_cell_value(c) for c in row])
        for summary in spec.get("summary_rows") or []:
            ws.append([_cell_value(c) for c in summary])
        _auto_column_width(ws, headers, rows)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value)


def send_excel(filename: str, sheets: Sequence[dict[str, Any]]):
    data = build_excel_workbook(sheets)
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
